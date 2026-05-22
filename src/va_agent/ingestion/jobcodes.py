"""Ingest the VA Duty MOS Noise Exposure Listing into the Job Code spine.

Reads the Fast Letter 10-35 xlsx (committed at ``data/duty_mos_noise.xlsx``)
and MERGEs one ``:CFR:JobCode {code, title, branch, type}`` node per row,
plus a ``:NOISE_EXPOSURE {probability}`` edge to the ``hearing`` Anatomy
node. The result is the authoritative spine: every job code in every branch
is present, with the VA's official noise-exposure probability attached.

Each sheet in the workbook has a different column layout, so this module
holds a per-sheet parser dispatch. Unparseable rows are dropped into
``data/review_queue.jsonl`` with a reason.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from ..graph.driver import GraphDriver
from .cfr import review_queue_path

# Canonical probability labels used downstream.
_PROB_HIGH = "Highly Probable"
_PROB_MOD = "Moderate"
_PROB_LOW = "Low"


@dataclass
class JobCodeRow:
    """One validated row, ready to write."""

    code: str
    title: str
    branch: str  # e.g. "Army", "Marine Corps", "Navy", "Air Force", "Coast Guard"
    type: str  # "MOS", "AFSC", "Navy Rating", "NEC"
    probability: str  # one of _PROB_HIGH / _PROB_MOD / _PROB_LOW


@dataclass
class IngestionReport:
    """What happened during one spine ingestion call."""

    rows_seen: int = 0
    rows_written: int = 0
    rows_rejected: list[tuple[str, str]] = field(default_factory=list)
    per_branch: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


# --- branch / sheet metadata --------------------------------------------------

# Maps sheet name -> (branch, type)
_SHEET_BRANCH: dict[str, tuple[str, str]] = {
    "ARMY ENLISTED": ("Army", "MOS"),
    "ARMY OFFICER": ("Army", "MOS"),
    "NAVY ENLISTED": ("Navy", "Navy Rating"),
    "NAVY OFFICER": ("Navy", "Navy Rating"),
    "MARINE CORPS": ("Marine Corps", "MOS"),
    "AIR FORCE ENLISTED": ("Air Force", "AFSC"),
    "AIR FORCE OFFICERS": ("Air Force", "AFSC"),
    "COAST GUARD": ("Coast Guard", "Navy Rating"),
    "MERCHANT MARINE": ("Merchant Marine", "Navy Rating"),
}


def _normalise_probability(high: Any, mod: Any, low: Any) -> str | None:
    """The xlsx encodes probability as an 'X' (or sometimes ' ') in one of three
    columns. Return the canonical label, or None if no column is marked."""

    def marked(v: Any) -> bool:
        # Most rows mark with 'X', but a handful of Army Officer / Air Force
        # Enlisted entries use 'Y'. Both are treated as the same flag.
        return isinstance(v, str) and v.strip().upper() in {"X", "Y"}

    # Highly Probable wins if multiple are marked (the xlsx occasionally has
    # whitespace artefacts in the other columns).
    if marked(high):
        return _PROB_HIGH
    if marked(mod):
        return _PROB_MOD
    if marked(low):
        return _PROB_LOW
    return None


# --- per-sheet parsers --------------------------------------------------------


def _parse_army(rows: Iterable[tuple], branch: str, type_: str) -> Iterable[JobCodeRow | tuple[str, dict]]:
    """ARMY ENLISTED / ARMY OFFICER: MOS, JOB TITLE, HIGHLY, MOD, LOW."""
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        code = row[0]
        title = row[1]
        if code in (None, "MOS"):
            continue
        if not isinstance(code, str) or not code.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        prob = _normalise_probability(row[2], row[3], row[4])
        if prob is None:
            yield ("missing_probability", {"code": code, "row": list(row)})
            continue
        yield JobCodeRow(
            code=code.strip(),
            title=(title or "").strip() if isinstance(title, str) else str(title or "").strip(),
            branch=branch,
            type=type_,
            probability=prob,
        )


def _parse_navy_enlisted(rows: Iterable[tuple], branch: str, type_: str):
    """NAVY ENLISTED: SERIES, RATING, JOB TITLE, HIGHLY, MOD, LOW."""
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        rating = row[1]
        title = row[2]
        if rating in (None, "RATING"):
            continue
        if not isinstance(rating, str) or not rating.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        prob = _normalise_probability(row[3], row[4], row[5])
        if prob is None:
            yield ("missing_probability", {"code": rating, "row": list(row)})
            continue
        yield JobCodeRow(
            code=rating.strip(),
            title=(title or "").strip() if isinstance(title, str) else str(title or "").strip(),
            branch=branch,
            type=type_,
            probability=prob,
        )


def _parse_navy_officer(rows: Iterable[tuple], branch: str, type_: str):
    """NAVY OFFICER: FIELD, GROUP, HIGHLY, MOD, LOW.

    The 'code' here is the leading designator-range from the GROUP column,
    e.g. ``0000-0099``. Title is the rest of the group label.
    """
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        group = row[1]
        if group in (None, "GROUP"):
            continue
        if not isinstance(group, str) or not group.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        # Split "0000-0099 HEALTH SERVICES MANAGEMENT GROUP" -> code + title.
        parts = group.strip().split(None, 1)
        code = parts[0]
        title = parts[1] if len(parts) > 1 else ""
        prob = _normalise_probability(row[2], row[3], row[4])
        if prob is None:
            yield ("missing_probability", {"code": code, "row": list(row)})
            continue
        yield JobCodeRow(
            code=code,
            title=title.strip(),
            branch=branch,
            type=type_,
            probability=prob,
        )


def _parse_marine(rows: Iterable[tuple], branch: str, type_: str):
    """MARINE CORPS: GRADE, MOS, OCCUPATIONAL FIELD, HIGHLY, MOD, LOW, ...

    Marine Corps lumps MOSs by Occupational Field (01XX, 02XX, ...) and splits
    by grade (O/W/E). The composite key code is ``{MOS}-{GRADE}``.
    """
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        grade = row[0]
        mos = row[1]
        title = row[2]
        if mos in (None, "MOS"):
            continue
        if grade not in ("O", "W", "E"):
            # Skip header/footnote rows that have no grade.
            continue
        if not isinstance(mos, str) or not mos.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        prob = _normalise_probability(row[3], row[4], row[5])
        if prob is None:
            yield ("missing_probability", {"code": mos, "row": list(row)})
            continue
        yield JobCodeRow(
            code=f"{mos.strip()}-{grade}",
            title=(title or "").strip() if isinstance(title, str) else str(title or "").strip(),
            branch=branch,
            type=type_,
            probability=prob,
        )


def _parse_af_enlisted(rows: Iterable[tuple], branch: str, type_: str):
    """AIR FORCE ENLISTED: 'AFSC -- JOB TITLE', HIGHLY, MOD, LOW."""
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        combined = row[0]
        if combined in (None, "AFSC - JOB TITLE"):
            continue
        if not isinstance(combined, str) or not combined.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        # Split on '--' or '-' (the spreadsheet uses double-dash but tolerate single).
        sep = "--" if "--" in combined else "-"
        parts = combined.split(sep, 1)
        code = parts[0].strip()
        title = parts[1].strip() if len(parts) > 1 else ""
        prob = _normalise_probability(row[1], row[2], row[3])
        if prob is None:
            yield ("missing_probability", {"code": code, "row": list(row)})
            continue
        yield JobCodeRow(
            code=code,
            title=title,
            branch=branch,
            type=type_,
            probability=prob,
        )


def _parse_af_officer(rows: Iterable[tuple], branch: str, type_: str):
    """AIR FORCE OFFICERS: AFSC, JOB TITLE, HIGHLY, MOD, LOW (+ trailing None)."""
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        code = row[0]
        title = row[1]
        if code in (None, "AFSC", "AFSC "):
            continue
        if isinstance(code, str) and code.strip() in ("AFSC",):
            continue
        if not isinstance(code, str) or not code.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        prob = _normalise_probability(row[2], row[3], row[4])
        if prob is None:
            yield ("missing_probability", {"code": code, "row": list(row)})
            continue
        yield JobCodeRow(
            code=code.strip(),
            title=(title or "").strip() if isinstance(title, str) else str(title or "").strip(),
            branch=branch,
            type=type_,
            probability=prob,
        )


def _parse_coast_guard(rows: Iterable[tuple], branch: str, type_: str):
    """COAST GUARD: RATINGS, JOB TITLE, HIGHLY, MOD, LOW."""
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        code = row[0]
        title = row[1]
        if code in (None, "RATINGS", "RATING"):
            continue
        if not isinstance(code, str) or not code.strip():
            yield ("invalid_code", {"row": list(row)})
            continue
        prob = _normalise_probability(row[2], row[3], row[4])
        if prob is None:
            yield ("missing_probability", {"code": code, "row": list(row)})
            continue
        yield JobCodeRow(
            code=code.strip(),
            title=(title or "").strip() if isinstance(title, str) else str(title or "").strip(),
            branch=branch,
            type=type_,
            probability=prob,
        )


_SHEET_PARSERS = {
    "ARMY ENLISTED": _parse_army,
    "ARMY OFFICER": _parse_army,
    "NAVY ENLISTED": _parse_navy_enlisted,
    "NAVY OFFICER": _parse_navy_officer,
    "MARINE CORPS": _parse_marine,
    "AIR FORCE ENLISTED": _parse_af_enlisted,
    "AIR FORCE OFFICERS": _parse_af_officer,
    "COAST GUARD": _parse_coast_guard,
    "MERCHANT MARINE": _parse_coast_guard,  # same column layout
}


# --- graph writes -------------------------------------------------------------


def _write_job_code(driver: GraphDriver, row: JobCodeRow) -> None:
    driver.cfr_write(
        """
        MERGE (a:CFR:Anatomy {name: $anatomy_name})
          SET a.body_system = $body_system,
              a.side = coalesce(a.side, $side)
        MERGE (jc:CFR:JobCode {code: $code, branch: $branch})
          SET jc.title = $title,
              jc.type  = $type
        MERGE (jc)-[r:NOISE_EXPOSURE]->(a)
          SET r.probability = $probability
        """,
        anatomy_name="hearing",
        body_system="hearing",
        side="unspecified",
        code=row.code,
        branch=row.branch,
        title=row.title,
        type=row.type,
        probability=row.probability,
    )


def _record_reject(path: Path, sheet: str, reason: str, payload: dict) -> None:
    entry = {"source": "jobcode_spine", "sheet": sheet, "reason": reason, **payload}
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, default=str) + "\n")


# --- public entry point -------------------------------------------------------


def ingest_job_code_spine(
    xlsx_path: Path | str,
    driver: GraphDriver,
    *,
    project_root: Path | None = None,
    sheets: list[str] | None = None,
) -> IngestionReport:
    """Read the Duty MOS Noise Exposure xlsx and build the JobCode spine.

    Args:
        xlsx_path: Path to the workbook.
        driver: GraphDriver wrapping a live Neo4j instance.
        project_root: Override the project root used for the review queue path.
        sheets: Optional restricted set of sheet names (used by tests).
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)
    review_path = review_queue_path(project_root)
    report = IngestionReport()

    sheet_names = sheets if sheets is not None else list(_SHEET_PARSERS.keys())
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            report.warnings.append(f"sheet not in workbook: {sheet_name}")
            continue
        ws = wb[sheet_name]
        branch, type_ = _SHEET_BRANCH[sheet_name]
        parser = _SHEET_PARSERS[sheet_name]
        for item in parser(ws.iter_rows(values_only=True), branch, type_):
            report.rows_seen += 1
            if isinstance(item, tuple):
                reason, payload = item
                _record_reject(review_path, sheet_name, reason, payload)
                report.rows_rejected.append((sheet_name, reason))
                continue
            _write_job_code(driver, item)
            report.rows_written += 1
            report.per_branch[branch] = report.per_branch.get(branch, 0) + 1

    return report


__all__ = [
    "IngestionReport",
    "JobCodeRow",
    "ingest_job_code_spine",
]
